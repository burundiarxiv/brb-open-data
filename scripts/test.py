import csv
from email.header import Header
from operator import ge
from bs4 import BeautifulSoup
import requests
from urllib.parse import unquote
import pyexcel as converter
import openpyxl
import pandas as pd


URL = "https://brb.bi" # The base link 

#This function is used to gather all the links that we will need. It just needs a path to know the path of the page
def get_everything(path): 
    page = requests.get(URL + path)
    soup = BeautifulSoup(page.content, "html.parser")
    results = soup.find_all("div",class_= "section")[2] # The second occurence of the section class is what contains our section
    others = results.find_all("p") # This contains all the links  
    links = [] 
    for i in others: #Get all the links
        links.append(URL + i.find_all("a")[0]["href"])
    return links #Then return them

def download_x(element): #This function downloads a specified element and converts the xls file to xlsx
    links = get_everything("/fr/content/balance-des-paiements") #Our path
    for link in links: 
        if element in unquote(link):
            data = requests.get(link) #Get the binary from the page
            open(unquote("../xls/"+link[35:]),'wb').write(data.content) #Keep the file as it is. And save it on the device. 
            converter.save_book_as(file_name="../xls/"+unquote(link[35:]), dest_file_name="../xls/"+unquote(link[35:])+"x") # Using a converter to convert the xls to xlsx
            print("Successfully Created ",unquote(link[35:])+"x")
            return "../xls/"+unquote(link[35:])+"x" #Return the name of the file

def remove_sheets(workbook, sheet_to_keep): #Remove unwanted sheets
    workbook_copy = workbook
    for sheet in workbook.sheetnames:
        if sheet != sheet_to_keep:
            del workbook_copy[sheet]
    return workbook_copy

def remove_n_lines(workbook, start, end): #remove problematic lines
    workbook_copy = workbook
    sheet = workbook_copy["Mensuelle"]
    sheet.delete_rows(start,end)
    return workbook_copy

def xlsx_to_csv(workbook_path):
    read_file = pd.read_excel(workbook_path)
    read_file.to_csv("../csv/Trial.csv", index=None, header= True)

def create_workbook_from_net(online_notebook): #main function, may need changes
    book = openpyxl.load_workbook(download_x(online_notebook))
    book = remove_sheets(book,"Mensuelle")
    book = remove_n_lines(book,0,7)
    book = remove_n_lines(book,99,300)
    book.save("../xls/trial.xlsx")
    print(book.sheetnames)

    xlsx_to_csv("../xls/trial.xlsx")

create_workbook_from_net("Importations_par_rubrique en valeur")
