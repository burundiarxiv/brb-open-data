import csv
from email.header import Header
from operator import ge
from bs4 import BeautifulSoup
import requests
from urllib.parse import unquote
import pyexcel as converter
import openpyxl
import pandas as pd


URL = "https://brb.bi"

def get_everything(path):
    page = requests.get(URL + path)
    soup = BeautifulSoup(page.content, "html.parser")
    results = soup.find_all("div",class_= "section")[2] # The second occurence of the section class is what contains our section
    others = results.find_all("p")
    links = [] 
    for i in others:
        links.append(URL + i.find_all("a")[0]["href"])
    return links
def download_x(element):
    links = get_everything("/fr/content/balance-des-paiements")
    for link in links:
        if element in unquote(link):
            data = requests.get(link)
            open(unquote(link[35:]),'wb').write(data.content)
            converter.save_book_as(file_name=unquote(link[35:]), dest_file_name=unquote(link[35:])+"x")
            print("Successfully Created ",unquote(link[35:]) )
            return unquote(link[35:])+"x"

def remove_sheets(workbook, sheet_to_keep):
    workbook_copy = workbook
    for sheet in workbook.sheetnames:
        if sheet != sheet_to_keep:
            del workbook_copy[sheet]
    return workbook_copy

def remove_n_lines(workbook, start, end):
    workbook_copy = workbook
    sheet = workbook_copy["Mensuelle"]
    sheet.delete_rows(start,end)
    return workbook_copy
def xlsx_to_csv(workbook_path):
    read_file = pd.read_excel(workbook_path)
    read_file.to_csv("Trial.csv", index=None, header= True)

def create_workbook_from_net(online_notebook):
    book = openpyxl.load_workbook(download_x(online_notebook))
    book = remove_sheets(book,"Mensuelle")
    book = remove_n_lines(book,0,7)
    book = remove_n_lines(book,99,300)
    book.save("trial.xlsx")
    print(book.sheetnames)

    xlsx_to_csv("trial.xlsx")

create_workbook_from_net("Importations_par_rubrique en valeur")